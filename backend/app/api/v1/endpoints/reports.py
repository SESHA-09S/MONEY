"""Report generation endpoints — PDF, Excel, CSV."""
from datetime import date
from io import BytesIO
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_active_user, get_db
from app.models.expense import Expense
from app.models.income import Income
from app.models.user import User
from app.services.business_service import BusinessService

router = APIRouter(prefix="/reports", tags=["Reports"])


async def _get_business(business_id: int, user: User, db):
    svc = BusinessService(db)
    biz = await svc.get_by_id(business_id)
    if not biz or (biz.owner_id != user.id and user.role.value != "admin"):
        raise HTTPException(status_code=404, detail="Business not found")
    return biz


@router.get("/{business_id}/export")
async def export_report(
    business_id: int,
    format: Literal["pdf", "excel", "csv"] = Query("pdf"),
    report_type: Literal["income", "expense", "cashflow", "summary"] = Query("summary"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Export financial reports in PDF, Excel, or CSV format."""
    business = await _get_business(business_id, current_user, db)

    inc_result = await db.execute(
        select(Income).where(Income.business_id == business_id)
    )
    exp_result = await db.execute(
        select(Expense).where(Expense.business_id == business_id)
    )
    incomes = inc_result.scalars().all()
    expenses = exp_result.scalars().all()

    if start_date:
        incomes = [i for i in incomes if i.transaction_date >= start_date]
        expenses = [e for e in expenses if e.transaction_date >= start_date]
    if end_date:
        incomes = [i for i in incomes if i.transaction_date <= end_date]
        expenses = [e for e in expenses if e.transaction_date <= end_date]

    if format == "csv":
        return _generate_csv(incomes, expenses, report_type, business.name)
    elif format == "excel":
        return _generate_excel(incomes, expenses, report_type, business.name)
    else:
        return _generate_pdf(incomes, expenses, report_type, business.name)


def _generate_csv(incomes, expenses, report_type, business_name):
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)

    if report_type in ("income", "summary"):
        writer.writerow(["Date", "Category", "Amount", "Payment Method", "Description"])
        for inc in sorted(incomes, key=lambda x: x.transaction_date, reverse=True):
            writer.writerow([inc.transaction_date, inc.category.value, inc.amount,
                             inc.payment_method.value, inc.description or ""])

    if report_type in ("expense", "summary"):
        if report_type == "summary":
            writer.writerow([])
            writer.writerow(["--- EXPENSES ---"])
        writer.writerow(["Date", "Category", "Amount", "Vendor", "Description"])
        for exp in sorted(expenses, key=lambda x: x.transaction_date, reverse=True):
            writer.writerow([exp.transaction_date, exp.category.value, exp.amount,
                             exp.vendor_name or "", exp.description or ""])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={business_name}_report.csv"},
    )


def _generate_excel(incomes, expenses, report_type, business_name):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()

    # Income sheet
    ws_inc = wb.active
    ws_inc.title = "Income"
    headers = ["Date", "Category", "Amount", "Payment Method", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws_inc.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.font = Font(bold=True, color="FFFFFF")
    for row, inc in enumerate(sorted(incomes, key=lambda x: x.transaction_date, reverse=True), 2):
        ws_inc.cell(row=row, column=1, value=str(inc.transaction_date))
        ws_inc.cell(row=row, column=2, value=inc.category.value)
        ws_inc.cell(row=row, column=3, value=inc.amount)
        ws_inc.cell(row=row, column=4, value=inc.payment_method.value)
        ws_inc.cell(row=row, column=5, value=inc.description or "")

    # Expense sheet
    ws_exp = wb.create_sheet("Expenses")
    for col, h in enumerate(["Date", "Category", "Amount", "Vendor", "Description", "Anomaly"], 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="EF4444")
    for row, exp in enumerate(sorted(expenses, key=lambda x: x.transaction_date, reverse=True), 2):
        ws_exp.cell(row=row, column=1, value=str(exp.transaction_date))
        ws_exp.cell(row=row, column=2, value=exp.category.value)
        ws_exp.cell(row=row, column=3, value=exp.amount)
        ws_exp.cell(row=row, column=4, value=exp.vendor_name or "")
        ws_exp.cell(row=row, column=5, value=exp.description or "")
        ws_exp.cell(row=row, column=6, value="Yes" if exp.is_anomaly else "No")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={business_name}_report.xlsx"},
    )


def _generate_pdf(incomes, expenses, report_type, business_name):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"SmartCash AI — Financial Report", styles["Title"]))
    elements.append(Paragraph(f"Business: {business_name}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)

    summary_data = [
        ["Metric", "Value"],
        ["Total Income", f"₹{total_income:,.2f}"],
        ["Total Expense", f"₹{total_expense:,.2f}"],
        ["Net Profit", f"₹{total_income - total_expense:,.2f}"],
        ["Transactions", str(len(incomes) + len(expenses))],
    ]
    t = Table(summary_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={business_name}_report.pdf"},
    )
